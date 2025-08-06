"""
Módulo para exportar rutinas de entrenamiento a Excel
Genera archivos Excel simples y profesionales siguiendo el formato estándar
"""

import pandas as pd
import logging
import re
from datetime import datetime
from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.athlete_manager import get_athlete_data
from modules.email_manager import show_email_sending_interface

def create_simple_routine_excel(athlete_id, routine_text):
    """Crea un Excel simple y limpio con el formato estándar usado por los entrenadores"""
    try:
        # Obtener datos del atleta
        athlete_data = get_athlete_data(athlete_id)
        if not athlete_data:
            logging.error(f"No se encontraron datos del atleta {athlete_id}")
            return None

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan de Entrenamiento"

        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Información del atleta (encabezado)
        ws.cell(row=1, column=1, value="ATLETA:")
        ws.cell(row=1, column=2, value=athlete_data['name'])
        ws.cell(row=2, column=1, value="DEPORTE:")
        ws.cell(row=2, column=2, value=athlete_data['sport'])
        ws.cell(row=3, column=1, value="NIVEL:")
        ws.cell(row=3, column=2, value=athlete_data['level'])
        ws.cell(row=4, column=1, value="FECHA:")
        ws.cell(row=4, column=2, value=datetime.now().strftime("%d/%m/%Y"))

        # Configurar estilos del encabezado
        for row in range(1, 5):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Parsear rutina de forma simple
        routine_data = parse_routine_simple(routine_text)
        
        # Empezar el plan desde la fila 6
        current_row = 6
        
        # Headers de la tabla
        headers = ["EJERCICIO", "SERIES/REPETICIONES", "CARGA", "NOTAS"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        current_row += 1

        # Agregar días y ejercicios
        for day_info in routine_data:
            # Header del día
            day_cell = ws.cell(row=current_row, column=1)
            day_cell.value = f"DÍA {day_info['day']} - {day_info['title']}"
            day_cell.font = day_font
            day_cell.fill = day_fill
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border = border
            
            # Merge células para el día
            ws.merge_cells(start_row=current_row, start_column=1, 
                          end_row=current_row, end_column=4)
            
            current_row += 1

            # Ejercicios del día
            for exercise in day_info['exercises']:
                if exercise['name']:  # Solo si hay nombre de ejercicio
                    ws.cell(row=current_row, column=1, value=exercise['name']).border = border
                    ws.cell(row=current_row, column=2, value=exercise['sets_reps']).border = border
                    ws.cell(row=current_row, column=3, value="").border = border  # Carga vacía para llenar
                    ws.cell(row=current_row, column=4, value=exercise['notes']).border = border
                    current_row += 1

            # Espacio entre días
            current_row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 45  # Ejercicio
        ws.column_dimensions['B'].width = 20  # Series/Repeticiones
        ws.column_dimensions['C'].width = 15  # Carga
        ws.column_dimensions['D'].width = 30  # Notas

        # Guardar en BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logging.info(f"Excel simple generado exitosamente para atleta {athlete_id}")
        return excel_buffer.getvalue()

    except Exception as e:
        logging.error(f"Error al crear Excel simple: {e}")
        return None

def parse_routine_simple(routine_text):
    """Parsea el texto de rutina respetando la estructura exacta por días - TODOS los bloques separados"""
    try:
        # Limpiar texto
        routine_text = routine_text.replace("[INICIO_NUEVA_RUTINA]", "").strip()
        
        days = []
        lines = routine_text.split('\n')
        
        current_day = None
        current_exercises = []
        inside_day = False
        
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Saltar líneas vacías
            if not line:
                i += 1
                continue
            
            # Detectar inicio de día/sesión (SESIÓN X, DÍA X, ### SESIÓN X)
            day_patterns = [
                r'###\s*(sesión|día)\s+(\d+)',  # ### SESIÓN 1
                r'^(sesión|día)\s+(\d+)',       # SESIÓN 1
                r'^\*\*\*seSIÓN\s+(\d+)', # ***SESIÓN 1
            ]
            
            day_found = False
            for pattern in day_patterns:
                day_match = re.search(pattern, line, re.IGNORECASE)
                if day_match:
                    day_found = True
                    
                    # Guardar día anterior si existe
                    if current_day and current_exercises:
                        days.append({
                            'day': current_day['day'],
                            'title': current_day['title'],
                            'exercises': current_exercises
                        })
                    
                    # Extraer número de día y título
                    if len(day_match.groups()) == 2:
                        day_num = day_match.group(2)
                    else:
                        day_num = day_match.group(3)
                    
                    # Extraer título después del guión
                    title = line.split('-')[-1].strip() if '-' in line else "ENTRENAMIENTO"
                    title = re.sub(r'###|\*\*\*', '', title).strip()
                    
                    current_day = {'day': day_num, 'title': title}
                    current_exercises = []
                    inside_day = True
                    break
            
            if day_found:
                i += 1
                continue
            
            # Solo procesar si estamos dentro de un día
            if not inside_day:
                i += 1
                continue
            
            # Detectar si la siguiente línea es otro día (para no procesar contenido fuera de contexto)
            is_next_day = False
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                for pattern in day_patterns:
                    if re.search(pattern, next_line, re.IGNORECASE):
                        is_next_day = True
                        break
            
            # Detectar TODOS los tipos de bloques/secciones
            section_keywords = [
                'BLOQUE', 'ACTIVACIÓN', 'POTENCIA', 'FUERZA', 'CONTRASTE', 
                'CIRCUITO', 'CALENTAMIENTO', 'CORE', 'ESTABILIDAD', 
                'VELOCIDAD', 'AGILIDAD', 'PLIOMETRÍA', 'TÉCNICO',
                'VUELTA A LA CALMA', 'ESTIRAMIENTOS', 'MOVILIDAD'
            ]
            
            is_section = any(keyword in line.upper() for keyword in section_keywords)
            
            if is_section:
                # Es una sección - agregar como separador visual
                current_exercises.append({
                    'name': f"** {line.upper()} **",
                    'sets_reps': '',
                    'notes': ''
                })
                i += 1
                continue
            
            # Detectar si es un ejercicio
            is_exercise = (
                line.startswith('-') or 
                line.startswith('•') or 
                line.startswith('*') or
                line.startswith('–') or  # guión largo
                re.search(r'\d+x\d+|\d+\s*rep|\(\d+|\d+\s*series', line.lower())
            )
            
            if is_exercise:
                # Limpiar prefijos
                cleaned_line = line
                for prefix in ['-', '•', '*', '–']:
                    if cleaned_line.startswith(prefix):
                        cleaned_line = cleaned_line[1:].strip()
                        break
                
                # Separar ejercicio de series/repeticiones
                exercise_name = cleaned_line
                sets_reps = ""
                notes = ""
                
                # Patrones más amplios para detectar series/repeticiones
                patterns = [
                    r'\((\d+x\d+/\d+)\)',     # (2x15/15)
                    r'\((\d+x\d+)\)',         # (3x10)
                    r'(\d+x\d+/\d+)',         # 2x15/15
                    r'(\d+x\d+)',             # 3x10
                    r'\((\d+)\s*rep\)',       # (15 rep)
                    r'(\d+)\s*rep',           # 15 rep
                    r'(\d+)\s*series?\s*de?\s*(\d+)',  # 3 series de 10
                    r'(\d+)\s*×\s*(\d+)',     # 3×10
                    r'\((\d+)\s*seg\)',       # (30 seg)
                    r'(\d+)\s*seg',           # 30 seg
                ]
                
                for pattern in patterns:
                    match = re.search(pattern, cleaned_line)
                    if match:
                        if len(match.groups()) == 1:
                            sets_reps = match.group(1)
                        else:
                            sets_reps = f"{match.group(1)}x{match.group(2)}"
                        exercise_name = re.sub(pattern, '', cleaned_line).strip()
                        break
                
                # Procesar información después de los dos puntos
                if ':' in exercise_name:
                    parts = exercise_name.split(':', 1)
                    exercise_name = parts[0].strip()
                    if not sets_reps and len(parts) > 1:
                        additional_info = parts[1].strip()
                        # Si contiene números, probablemente son series/reps
                        if re.search(r'\d+', additional_info):
                            # Verificar si es información de series/reps
                            if any(word in additional_info.lower() for word in ['rep', 'series', 'x', 'seg']):
                                sets_reps = additional_info
                            else:
                                notes = additional_info
                        else:
                            notes = additional_info
                
                # Limpiar nombre final
                exercise_name = exercise_name.rstrip('.').strip()
                
                # Solo agregar si hay nombre de ejercicio válido
                if exercise_name and len(exercise_name) > 2:
                    current_exercises.append({
                        'name': exercise_name,
                        'sets_reps': sets_reps,
                        'notes': notes
                    })
            
            i += 1
        
        # Agregar último día
        if current_day and current_exercises:
            days.append({
                'day': current_day['day'],
                'title': current_day['title'],
                'exercises': current_exercises
            })
        
        # Si no se encontraron días, crear uno general
        if not days:
            exercises = []
            for line in lines:
                line = line.strip()
                if (line and 
                    not any(keyword in line.upper() for keyword in ['METODOLOGÍA', 'PLAN', 'OBJETIVO']) and
                    len(line) > 3):
                    exercises.append({
                        'name': line,
                        'sets_reps': '',
                        'notes': ''
                    })
            
            if exercises:
                days = [{
                    'day': '1',
                    'title': 'ENTRENAMIENTO',
                    'exercises': exercises
                }]
        
        return days

    except Exception as e:
        logging.error(f"Error al parsear rutina simple: {e}")
        return [{
            'day': '1',
            'title': 'ENTRENAMIENTO',
            'exercises': [{'name': 'Error al procesar rutina', 'sets_reps': '', 'notes': ''}]
        }]

def generate_routine_excel_from_chat(athlete_id, chat_message):
    """Función principal para generar Excel desde mensaje del chat"""
    try:
        if not athlete_id or not chat_message:
            logging.error("athlete_id y chat_message son requeridos")
            return None, None

        # Obtener datos del atleta
        athlete_data = get_athlete_data(athlete_id)
        if not athlete_data:
            logging.error(f"No se encontraron datos del atleta {athlete_id}")
            return None, None

        athlete_name = athlete_data['name']
        
        # Generar Excel simple
        excel_data = create_simple_routine_excel(athlete_id, chat_message)
        
        if excel_data:
            logging.info(f"Excel generado exitosamente para {athlete_name}")
            
            # 💾 NUEVO: Guardar Excel en session_state para reutilización
            timestamp = int(datetime.now().timestamp())
            excel_key = f"excel_data_{athlete_id}_{timestamp}"
            filename_key = f"{excel_key}_filename"
            timestamp_key = f"{excel_key}_timestamp"
            
            athlete_name_clean = athlete_name.replace(' ', '_')
            filename = f"Rutina_{athlete_name_clean}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            
            # Guardar en session_state
            st.session_state[excel_key] = excel_data
            st.session_state[filename_key] = filename
            st.session_state[timestamp_key] = timestamp
            
            logging.info(f"📁 Excel guardado en session: {excel_key}")
            
            return excel_data, athlete_name
        else:
            logging.error(f"Error al generar Excel para {athlete_name}")
            return None, None

    except Exception as e:
        logging.error(f"Error en generate_routine_excel_from_chat: {e}")
        return None, None

def create_download_button(excel_data, athlete_name, filename_prefix="Rutina", unique_id=""):
    """Crea un botón de descarga para el archivo Excel"""
    try:
        if not excel_data:
            return False

        # Crear nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{athlete_name.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Crear botón de descarga
        st.download_button(
            label="📥 Descargar Rutina en Excel",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_excel_{unique_id}_{timestamp}",
            use_container_width=True,
            type="primary"
        )
        
        return True

    except Exception as e:
        logging.error(f"Error al crear botón de descarga: {e}")
        return False

def create_download_and_email_interface(athlete_id, excel_data, athlete_name, filename_prefix="Rutina", unique_id=""):
    """Crea interfaz completa con descarga y envío por email"""
    try:
        if not excel_data:
            st.error("❌ No hay datos de Excel para procesar")
            return False

        # Crear nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{athlete_name.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Obtener datos del atleta
        athlete_data = get_athlete_data(athlete_id)
        if not athlete_data:
            st.error("❌ No se pudieron obtener los datos del atleta")
            return False

        # Crear dos columnas para descarga y email
        col1, col2 = st.columns(2)
        
        with col1:
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_excel_{unique_id}_{timestamp}",
                use_container_width=True,
                type="primary"
            )
        
        with col2:
            if st.button(
                "📧 Enviar por Email", 
                key=f"email_button_{unique_id}_{timestamp}",
                use_container_width=True,
                type="secondary"
            ):
                st.session_state[f'show_email_interface_{unique_id}'] = True
                st.rerun()
        
        # Mostrar interfaz de email si se solicitó
        if st.session_state.get(f'show_email_interface_{unique_id}', False):
            st.markdown("---")
            show_email_sending_interface(athlete_data, excel_data, filename)
            
            # Botón para ocultar interfaz de email
            if st.button("❌ Cancelar envío", key=f"cancel_email_{unique_id}"):
                st.session_state[f'show_email_interface_{unique_id}'] = False
                st.rerun()
        
        return True

    except Exception as e:
        st.error(f"Error en la interfaz: {e}")
        logging.error(f"Error al crear interfaz de descarga y email: {e}")
        return False
