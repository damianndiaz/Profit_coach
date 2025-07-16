"""
Módulo para exportar rutinas de entrenamiento a Excel
Genera archivos Excel simples y profesionales siguiendo el formato estándar
"""

import pandas as pd
import logging
import re
from datetime import datetime
from io import BytesIO
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.athlete_manager import get_athlete_data

def create_simple_routine_excel(athlete_id, routine_text):
    """Crea un Excel simple y limpio con el formato estándar usado por los entrenadores"""
    try:
        # Obtener datos del atleta
        athlete_data = get_athlete_data(athlete_id)
        if not athlete_data:
            logging.error(f"No se encontraron datos del atleta {athlete_id}")
            return None

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Plan de Entrenamiento"

        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF", size=11)
        day_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Información del atleta (encabezado)
        ws.cell(row=1, column=1, value="ATLETA:")
        ws.cell(row=1, column=2, value=athlete_data['name'])
        ws.cell(row=2, column=1, value="DEPORTE:")
        ws.cell(row=2, column=2, value=athlete_data['sport'])
        ws.cell(row=3, column=1, value="NIVEL:")
        ws.cell(row=3, column=2, value=athlete_data['level'])
        ws.cell(row=4, column=1, value="FECHA:")
        ws.cell(row=4, column=2, value=datetime.now().strftime("%d/%m/%Y"))

        # Configurar estilos del encabezado
        for row in range(1, 5):
            ws.cell(row=row, column=1).font = Font(bold=True)

        # Parsear rutina de forma simple
        routine_data = parse_routine_simple(routine_text)
        
        # Empezar el plan desde la fila 6
        current_row = 6
        
        # Headers de la tabla
        headers = ["EJERCICIO", "SERIES/REPETICIONES", "CARGA", "NOTAS"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        current_row += 1

        # Agregar días y ejercicios
        for day_info in routine_data:
            # Header del día
            day_cell = ws.cell(row=current_row, column=1)
            day_cell.value = f"DÍA {day_info['day']} - {day_info['title']}"
            day_cell.font = day_font
            day_cell.fill = day_fill
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border = border
            
            # Merge células para el día
            ws.merge_cells(start_row=current_row, start_column=1, 
                          end_row=current_row, end_column=4)
            
            current_row += 1

            # Ejercicios del día
            for exercise in day_info['exercises']:
                if exercise['name']:  # Solo si hay nombre de ejercicio
                    ws.cell(row=current_row, column=1, value=exercise['name']).border = border
                    ws.cell(row=current_row, column=2, value=exercise['sets_reps']).border = border
                    ws.cell(row=current_row, column=3, value="").border = border  # Carga vacía para llenar
                    ws.cell(row=current_row, column=4, value=exercise['notes']).border = border
                    current_row += 1

            # Espacio entre días
            current_row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 45  # Ejercicio
        ws.column_dimensions['B'].width = 20  # Series/Repeticiones
        ws.column_dimensions['C'].width = 15  # Carga
        ws.column_dimensions['D'].width = 30  # Notas

        # Guardar en BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logging.info(f"Excel simple generado exitosamente para atleta {athlete_id}")
        return excel_buffer.getvalue()

    except Exception as e:
        logging.error(f"Error al crear Excel simple: {e}")
        return None

def parse_routine_simple(routine_text):
    """Parsea el texto de rutina de forma simple y eficiente"""
    try:
        # Limpiar texto
        routine_text = routine_text.replace("[INICIO_NUEVA_RUTINA]", "").strip()
        
        days = []
        lines = routine_text.split('\n')
        
        current_day = None
        current_exercises = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Detectar inicio de día
            day_match = re.search(r'día\s+(\d+)', line, re.IGNORECASE)
            if day_match:
                # Guardar día anterior si existe
                if current_day:
                    days.append({
                        'day': current_day['day'],
                        'title': current_day['title'],
                        'exercises': current_exercises
                    })
                
                # Iniciar nuevo día
                day_num = day_match.group(1)
                title = line.split('-')[-1].strip() if '-' in line else "ENTRENAMIENTO"
                current_day = {'day': day_num, 'title': title}
                current_exercises = []
                continue
            
            # Detectar secciones (BLOQUE, ACTIVACIÓN, etc.)
            if any(keyword in line.upper() for keyword in 
                  ['BLOQUE', 'ACTIVACIÓN', 'POTENCIA', 'FUERZA', 'CONTRASTE', 'CIRCUITO']):
                # Es una sección - agregar como separador
                current_exercises.append({
                    'name': f"** {line.upper()} **",
                    'sets_reps': '',
                    'notes': ''
                })
                continue
            
            # Es un ejercicio normal
            if line.startswith('-') or line.startswith('•') or line.startswith('*'):
                line = line[1:].strip()
            
            # Separar ejercicio de series/repeticiones
            exercise_name = line
            sets_reps = ""
            notes = ""
            
            # Buscar patrones de series/repeticiones
            patterns = [
                r'(\d+)\s*x\s*(\d+)',  # 3x10
                r'(\d+)\s*series?\s*de?\s*(\d+)',  # 3 series de 10
                r'(\d+)\s*×\s*(\d+)',  # 3×10
            ]
            
            for pattern in patterns:
                match = re.search(pattern, line)
                if match:
                    sets_reps = f"{match.group(1)}x{match.group(2)}"
                    exercise_name = re.sub(pattern, '', line).strip()
                    break
            
            # Limpiar nombre del ejercicio
            if ':' in exercise_name:
                parts = exercise_name.split(':')
                exercise_name = parts[0].strip()
                if not sets_reps and len(parts) > 1:
                    sets_reps = parts[1].strip()
            
            current_exercises.append({
                'name': exercise_name,
                'sets_reps': sets_reps,
                'notes': notes
            })
        
        # Agregar último día
        if current_day:
            days.append({
                'day': current_day['day'],
                'title': current_day['title'],
                'exercises': current_exercises
            })
        
        # Si no se encontraron días, crear uno general
        if not days:
            exercises = []
            for line in lines:
                line = line.strip()
                if line and not any(keyword in line.upper() for keyword in ['METODOLOGÍA', 'BLOQUE']):
                    exercises.append({
                        'name': line,
                        'sets_reps': '',
                        'notes': ''
                    })
            
            days = [{
                'day': '1',
                'title': 'ENTRENAMIENTO',
                'exercises': exercises
            }]
        
        return days

    except Exception as e:
        logging.error(f"Error al parsear rutina simple: {e}")
        return [{
            'day': '1',
            'title': 'ENTRENAMIENTO',
            'exercises': [{'name': 'Error al procesar rutina', 'sets_reps': '', 'notes': ''}]
        }]

def generate_routine_excel_from_chat(athlete_id, chat_message):
    """Función principal para generar Excel desde mensaje del chat"""
    try:
        if not athlete_id or not chat_message:
            logging.error("athlete_id y chat_message son requeridos")
            return None, None

        # Obtener datos del atleta
        athlete_data = get_athlete_data(athlete_id)
        if not athlete_data:
            logging.error(f"No se encontraron datos del atleta {athlete_id}")
            return None, None

        athlete_name = athlete_data['name']
        
        # Generar Excel simple
        excel_data = create_simple_routine_excel(athlete_id, chat_message)
        
        if excel_data:
            logging.info(f"Excel generado exitosamente para {athlete_name}")
            return excel_data, athlete_name
        else:
            logging.error(f"Error al generar Excel para {athlete_name}")
            return None, None

    except Exception as e:
        logging.error(f"Error en generate_routine_excel_from_chat: {e}")
        return None, None

def create_download_button(excel_data, athlete_name, filename_prefix="Rutina", unique_id=""):
    """Crea un botón de descarga para el archivo Excel"""
    try:
        if not excel_data:
            return False

        # Crear nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{filename_prefix}_{athlete_name.replace(' ', '_')}_{timestamp}.xlsx"
        
        # Crear botón de descarga
        st.download_button(
            label="📥 Descargar Rutina en Excel",
            data=excel_data,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"download_excel_{unique_id}_{timestamp}",
            use_container_width=True,
            type="primary"
        )
        
        return True

    except Exception as e:
        logging.error(f"Error al crear botón de descarga: {e}")
        return False
