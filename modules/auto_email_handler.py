"""
Sistema automático de envío de emails mejorado
Integración completa con el chat de IA
"""

import re
import time
import logging
import streamlit as st
from typing import Tuple, Optional
from modules.email_manager import send_routine_email
from modules.athlete_manager import get_athlete_data

class AutoEmailHandler:
    """Maneja el envío automático de emails desde el chat"""
    
    def __init__(self):
        # Patrones mejorados para detectar comandos de email
        self.email_patterns = [
            # Español - comandos directos
            r'\b(envía|envía|manda|mandá|envíalo|envialo|mandalo|mándalo)\s*(por)?\s*(email|mail|correo)\b',
            r'\b(email|mail|correo)\s*(me|lo|la|esto|eso|la\s*rutina|el\s*plan)\b',
            r'\b(quiero|necesito|puedes|podrías|me)\s*(que\s*)?(lo\s*)?(envíes|envies|mandes|mándalo|enviarlo|mandarlo)\s*(por)?\s*(email|mail|correo)\b',
            
            # Comandos con "por favor"
            r'\b(por\s*favor|please)\s*(envía|envía|manda|send)\s*(por)?\s*(email|mail|correo)\b',
            
            # Inglés básico
            r'\b(send|email|mail)\s*(it|this|the\s*routine|the\s*plan)?\s*(by|via)?\s*(email|mail)\b',
            
            # Patrones específicos
            r'\b(via|por)\s*(email|mail|correo)\b',
            r'\b(al\s*mail|por\s*mail|por\s*correo|al\s*email)\b',
            
            # Comandos más naturales
            r'\b(compártelo|compartelo|envíaselo|enviaselo|mandaselo|mándaselo)\s*(por)?\s*(email|mail|correo)?\b',
            
            # NUEVOS patrones más específicos
            r'\b(me\s*la\s*podes\s*mandar|me\s*la\s*puedes\s*mandar|me\s*lo\s*podes\s*mandar|me\s*lo\s*puedes\s*mandar)\s*(por)?\s*(email|mail|correo)?\b',
            r'\b(mándamela|mandamela|envíamela|enviamela)\s*(por)?\s*(email|mail|correo)?\b',
            r'\bmail\b(?!\s*(de|from|@))',  # Detectar "mail" solo (no en direcciones de email)
        ]
        
    def detect_email_command(self, message: str) -> bool:
        """Detecta si el mensaje contiene un comando de email"""
        message_lower = message.lower().strip()
        
        for pattern in self.email_patterns:
            if re.search(pattern, message_lower):
                logging.info(f"✅ Email command detected with pattern: {pattern}")
                return True
        
        return False
    
    def extract_email_from_message(self, message: str) -> Optional[str]:
        """Extrae email del mensaje si está especificado"""
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        matches = re.findall(email_pattern, message)
        return matches[0] if matches else None
    
    def process_email_request(self, athlete_id: int, routine_text: str, user_message: str = "") -> Tuple[bool, str]:
        """Procesa una solicitud de envío de email"""
        try:
            logging.info(f"🔄 PROCESSING EMAIL REQUEST - Athlete: {athlete_id}, Routine length: {len(routine_text)}")
            
            # Verificar que es una rutina válida
            if not self._is_valid_routine(routine_text):
                logging.warning(f"⚠️ Invalid routine detected for athlete {athlete_id}")
                return False, "❌ No se detectó una rutina válida para enviar"
            
            logging.info(f"✅ Valid routine detected for athlete {athlete_id}")
            
            # Obtener datos del atleta
            athlete_data = get_athlete_data(athlete_id)
            if not athlete_data:
                logging.error(f"❌ No athlete data found for ID {athlete_id}")
                return False, "❌ No se encontraron datos del atleta"
            
            logging.info(f"✅ Athlete data retrieved: {athlete_data.get('name', 'Unknown')}")
            
            # Verificar email del atleta
            athlete_email = athlete_data.get('email', '').strip()
            
            # Intentar extraer email del mensaje del usuario
            message_email = self.extract_email_from_message(user_message)
            if message_email:
                athlete_email = message_email
                logging.info(f"📧 Email extraído del mensaje: {message_email}")
            
            if not athlete_email:
                return False, "❌ No se encontró email del atleta. Especifica un email o actualiza los datos del atleta."
            
            # Validar formato de email
            if not self._is_valid_email(athlete_email):
                return False, f"❌ Email inválido: {athlete_email}"
            
            # Generar Excel de la rutina
            excel_data, filename = self._generate_routine_excel(athlete_data, routine_text)
            if not excel_data:
                return False, "❌ Error generando archivo Excel de la rutina"
            
            # Enviar email
            success, message = send_routine_email(
                athlete_email=athlete_email,
                athlete_name=athlete_data['name'],
                excel_data=excel_data,
                filename=filename,
                trainer_name="ProFit Coach AI"
            )
            
            if success:
                logging.info(f"✅ Email enviado exitosamente a {athlete_email}")
                return True, f"✅ Rutina enviada exitosamente a {athlete_email}"
            else:
                logging.error(f"❌ Error enviando email: {message}")
                return False, f"❌ Error enviando email: {message}"
                
        except Exception as e:
            error_msg = f"Error procesando solicitud de email: {str(e)}"
            logging.error(error_msg)
            return False, f"❌ {error_msg}"
    
    def send_existing_excel(self, athlete_id: int, excel_data: bytes, filename: str, user_message: str = "") -> Tuple[bool, str]:
        """
        Envía un Excel ya generado por email - NUEVA FUNCIONALIDAD
        Esta función usa archivos Excel existentes en lugar de generar nuevos
        """
        try:
            logging.info(f"📧 ENVIANDO EXCEL EXISTENTE - Athlete: {athlete_id}, File: {filename}")
            
            # Obtener datos del atleta
            athlete_data = get_athlete_data(athlete_id)
            if not athlete_data:
                logging.error(f"❌ No athlete data found for ID {athlete_id}")
                return False, "❌ No se encontraron datos del atleta"
            
            logging.info(f"✅ Athlete data retrieved: {athlete_data.get('name', 'Unknown')}")
            
            # Verificar email del atleta con validación robusta
            athlete_email = athlete_data.get('email') or ''
            if athlete_email is None:
                athlete_email = ''
            athlete_email = str(athlete_email).strip()
            
            # Intentar extraer email del mensaje del usuario
            message_email = self.extract_email_from_message(user_message)
            if message_email:
                athlete_email = message_email
                logging.info(f"📧 Email extraído del mensaje: {message_email}")
            
            if not athlete_email:
                return False, "❌ No se encontró email del atleta. Especifica un email o actualiza los datos del atleta."
            
            # Validar formato de email
            if not self._is_valid_email(athlete_email):
                return False, f"❌ Email inválido: {athlete_email}"
            
            logging.info(f"📧 Enviando Excel existente ({len(excel_data)} bytes) a {athlete_email}")
            
            # Enviar email usando el Excel ya generado
            success, message = send_routine_email(
                athlete_email=athlete_email,
                athlete_name=athlete_data['name'],
                excel_data=excel_data,
                filename=filename,
                trainer_name="ProFit Coach AI"
            )
            
            if success:
                logging.info(f"✅ Excel existente enviado exitosamente a {athlete_email}")
                return True, f"✅ Rutina enviada exitosamente a {athlete_email}"
            else:
                logging.error(f"❌ Error enviando Excel existente: {message}")
                return False, f"❌ Error enviando email: {message}"
                
        except Exception as e:
            error_msg = f"Error enviando Excel existente: {str(e)}"
            logging.error(error_msg)
            return False, f"❌ {error_msg}"

    def _is_valid_routine(self, routine_text: str) -> bool:
        """Verifica si el texto contiene una rutina válida"""
        routine_indicators = [
            "[INICIO_NUEVA_RUTINA]",
            "BLOQUE",
            "ACTIVACIÓN",
            "DINÁMICO",
            "FUERZA",
            "CONTRASTE",
            "ejercicio",
            "series",
            "repeticiones"
        ]
        
        routine_text_lower = routine_text.lower()
        return any(indicator.lower() in routine_text_lower for indicator in routine_indicators)
    
    def _is_valid_email(self, email: str) -> bool:
        """Valida formato de email"""
        email_pattern = r'^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}$'
        return re.match(email_pattern, email) is not None
    
    def _generate_routine_excel(self, athlete_data: dict, routine_text: str) -> Tuple[Optional[bytes], str]:
        """Genera el archivo Excel de la rutina"""
        try:
            from datetime import datetime
            from modules.routine_export import create_simple_routine_excel
            
            # Generar nombre de archivo con validación
            athlete_name = athlete_data.get('name', 'Atleta_Desconocido')
            if athlete_name is None:
                athlete_name = 'Atleta_Desconocido'
            athlete_name = str(athlete_name).replace(' ', '_')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            filename = f"Rutina_{athlete_name}_{timestamp}.xlsx"
            
            # Usar la función existente de routine_export
            # Nota: Esta función espera athlete_id, pero necesitamos adaptarla
            excel_data = self._create_excel_from_routine_text(athlete_data, routine_text)
            
            return excel_data, filename
            
        except Exception as e:
            logging.error(f"Error generando Excel: {e}")
            return None, ""
    
    def _create_excel_from_routine_text(self, athlete_data: dict, routine_text: str) -> Optional[bytes]:
        """Crea Excel directamente desde el texto de la rutina"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from datetime import datetime
            from io import BytesIO
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("No se pudo crear worksheet activo")
                
            ws.title = "Plan de Entrenamiento"

            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            day_font = Font(bold=True, color="FFFFFF", size=11)
            day_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

            # Información del atleta (encabezado)
            ws.cell(row=1, column=1, value="ATLETA:")
            ws.cell(row=1, column=2, value=athlete_data['name'])
            ws.cell(row=2, column=1, value="DEPORTE:")
            ws.cell(row=2, column=2, value=athlete_data['sport'])
            ws.cell(row=3, column=1, value="NIVEL:")
            ws.cell(row=3, column=2, value=athlete_data['level'])
            ws.cell(row=4, column=1, value="FECHA:")
            ws.cell(row=4, column=2, value=datetime.now().strftime("%d/%m/%Y"))

            # Estilo para el encabezado
            for row in range(1, 5):
                ws.cell(row=row, column=1).font = header_font
                ws.cell(row=row, column=1).fill = header_fill
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2).border = border

            # Procesar contenido de la rutina
            current_row = 6
            lines = routine_text.split('\n')
            
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Detectar diferentes tipos de contenido
                if any(keyword in line.upper() for keyword in ['BLOQUE', 'DÍA', 'ACTIVACIÓN', 'DINÁMICO', 'FUERZA', 'CONTRASTE']):
                    # Es un encabezado de bloque o día
                    ws.cell(row=current_row, column=1, value=line)
                    ws.cell(row=current_row, column=1).font = day_font
                    ws.cell(row=current_row, column=1).fill = day_fill
                    ws.cell(row=current_row, column=1).border = border
                    ws.merge_cells(f'A{current_row}:D{current_row}')
                    current_row += 1
                    
                elif re.search(r'\d+\s*x\s*\d+', line) or 'series' in line.lower() or 'repeticiones' in line.lower():
                    # Es un ejercicio con series/reps
                    parts = line.split('-', 1) if '-' in line else [line]
                    exercise_name = parts[0].strip()
                    exercise_details = parts[1].strip() if len(parts) > 1 else ""
                    
                    ws.cell(row=current_row, column=1, value=exercise_name)
                    ws.cell(row=current_row, column=2, value=exercise_details)
                    
                    for col in range(1, 5):
                        ws.cell(row=current_row, column=col).border = border
                    
                    current_row += 1
                    
                else:
                    # Texto general
                    ws.cell(row=current_row, column=1, value=line)
                    ws.merge_cells(f'A{current_row}:D{current_row}')
                    ws.cell(row=current_row, column=1).border = border
                    current_row += 1

            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Guardar en BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logging.error(f"Error creando Excel desde texto: {e}")
            return None
    
    def setup_auto_email_in_session(self, athlete_id: int, routine_text: str, user_message: str = ""):
        """Configura el envío automático en la sesión de Streamlit"""
        try:
            # Detectar si hay comando de email en el mensaje
            if not self.detect_email_command(user_message):
                return
            
            # Configurar flag en session_state para mostrar UI de envío
            st.session_state[f'pending_email_{athlete_id}'] = {
                'routine_text': routine_text,
                'user_message': user_message,
                'timestamp': str(int(time.time())),
                'auto_detected': True
            }
            
            logging.info(f"📧 Email automático configurado para atleta {athlete_id}")
            
        except Exception as e:
            logging.error(f"Error configurando email automático: {e}")
    
    def show_auto_email_ui(self, athlete_id: int):
        """Muestra la UI para confirmar envío automático"""
        session_key = f'pending_email_{athlete_id}'
        
        if session_key not in st.session_state:
            return
        
        email_data = st.session_state[session_key]
        
        # Mostrar notificación de email detectado
        with st.container():
            st.info("📧 **Comando de email detectado**")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                if st.button("✅ Enviar Rutina por Email", type="primary", use_container_width=True):
                    success, message = self.process_email_request(
                        athlete_id=athlete_id,
                        routine_text=email_data['routine_text'],
                        user_message=email_data.get('user_message', '')
                    )
                    
                    if success:
                        st.success(message)
                    else:
                        st.error(message)
                    
                    # Limpiar flag
                    del st.session_state[session_key]
                    st.rerun()
            
            with col2:
                if st.button("❌ Cancelar", use_container_width=True):
                    del st.session_state[session_key]
                    st.rerun()

# Instancia global del handler
auto_email_handler = AutoEmailHandler()
